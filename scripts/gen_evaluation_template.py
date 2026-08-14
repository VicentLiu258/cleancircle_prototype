# -*- coding: utf-8 -*-
"""
生成「AI 排课推荐抽查评测模板」Excel 文件
运行：python scripts/gen_evaluation_template.py
输出：docs/AI排课推荐抽查评测模板.xlsx
"""
import os, datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

wb = Workbook()

# ─── 公用样式 ───
BLUE   = PatternFill("solid", fgColor="4472C4")
LTBLUE = PatternFill("solid", fgColor="D6E4F0")
GREEN  = PatternFill("solid", fgColor="E2EFDA")
YELLOW = PatternFill("solid", fgColor="FFF2CC")
RED    = PatternFill("solid", fgColor="FCE4EC")
GRAY   = PatternFill("solid", fgColor="F2F2F2")
WHITE  = PatternFill("solid", fgColor="FFFFFF")
HDR_FONT  = Font(name="微软雅黑", bold=True, color="FFFFFF", size=11)
SUB_FONT  = Font(name="微软雅黑", bold=True, size=10)
BODY_FONT = Font(name="微软雅黑", size=10)
THIN = Side(style="thin", color="BFBFBF")
BOX  = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

def style_header(ws, row, col_count, fill=BLUE, font=HDR_FONT):
    for c in range(1, col_count + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill
        cell.font = font
        cell.alignment = CENTER
        cell.border = BOX

def style_rows(ws, start_row, end_row, col_count, alt_fill=GRAY):
    for r in range(start_row, end_row + 1):
        for c in range(1, col_count + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = BODY_FONT
            cell.alignment = LEFT if c > 1 else CENTER
            cell.border = BOX
            if r % 2 == 0:
                cell.fill = alt_fill

def auto_width(ws, col_count, min_w=12, max_w=36):
    for c in range(1, col_count + 1):
        best = min_w
        for row in ws.iter_rows(min_col=c, max_col=c, values_only=False):
            for cell in row:
                if cell.value:
                    best = max(best, min(max_w, len(str(cell.value)) * 1.8 + 2))
        ws.column_dimensions[get_column_letter(c)].width = best


# ════════════════════════════════════════════════════════════════
# Sheet 1 ─ 抽查评测记录表（核心工作表，业务人员每次评测填写）
# ════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "1-抽查评测记录"
ws1.sheet_properties.tabColor = "4472C4"

cols1 = [
    ("序号",          8),
    ("抽查日期",      14),
    ("抽样方式",      16),
    ("用户样本ID",    14),
    ("问卷提交时间",  16),
    ("问卷核心标签",  28),
    ("AI推荐方案名称",22),
    ("推荐天数",      10),
    ("命中排课规则",  22),
    ("日均时长(min)", 14),
    ("安全禁忌校验",  14),
    ("内容多样性",    14),
    ("难度匹配度",    14),
    ("推荐可解释性",  14),
    ("综合评定",      14),
    ("Bad Case类型",  18),
    ("问题描述",      36),
    ("建议优化规则",  28),
    ("关联B12规则ID", 16),
    ("评测人",        12),
    ("备注",          24),
]

for c, (name, w) in enumerate(cols1, 1):
    ws1.cell(row=1, column=c, value=name)
    ws1.column_dimensions[get_column_letter(c)].width = w
style_header(ws1, 1, len(cols1))

# 填充 5 行示例数据
examples = [
    [1, "2026-08-14", "概率采样(5%)", "U-8921", "2026-08-14 10:15",
     "黄体期/大基数/目标舒缓/无器械", "舒缓减压·7天", 7,
     "R-04 周期安全 + R-09 低冲击", 18, "通过", "良好(78%)", "匹配", "清晰",
     "推荐准确", "", "", "", "", "张三", ""],
    [2, "2026-08-14", "高风险标签", "U-6102", "2026-08-13 16:42",
     "久坐/肩颈不适/产后/未复缩", "肩颈修复·5天", 5,
     "R-09 低冲击", 20, "1项遗漏", "偏低(55%)", "偏高", "部分缺失",
     "Bad Case", "安全禁忌遗漏", "产后用户被推荐卷腹类动作，缺少腹直肌分离硬性排除",
     "新增「产后安全禁忌」排除项", "R-07", "李四", "已提交B12规则优化"],
    [3, "2026-08-14", "用户负反馈", "U-7743", "2026-08-12 09:30",
     "新手/低冲击/20min目标", "新手塑形·21天", 21,
     "R-01 新手保护 + R-11 多样性", 22, "通过", "良好(82%)", "匹配", "清晰",
     "推荐准确", "", "", "", "", "张三", "用户首周手动替换1次，属正常范围"],
    [4, "2026-08-14", "规则兜底采样", "U-5509", "2026-08-14 08:05",
     "新用户/经期第2天", "经期特别照顾·3天", 3,
     "R-04 周期安全", 15, "通过", "良好(70%)", "匹配", "清晰",
     "推荐准确", "", "", "", "", "王五", "兜底池命中，内容偏少但安全"],
    [5, "2026-08-14", "概率采样(5%)", "U-4320", "2026-08-13 20:11",
     "进阶/塑形/有器械/哑铃", "进阶塑形·14天", 14,
     "R-05 难度递进 + R-11 多样性", 28, "通过", "偏低(48%)", "匹配", "部分缺失",
     "Bad Case", "内容过于重复", "连续3天推荐同系列哑铃视频，多样性权重被难度规则覆盖",
     "提升多样性规则优先级", "R-11", "李四", ""],
]
for r, row_data in enumerate(examples, 2):
    for c, val in enumerate(row_data, 1):
        ws1.cell(row=r, column=c, value=val)
style_rows(ws1, 2, 6, len(cols1))

# 下拉数据验证
dv_sample = DataValidation(type="list", formula1='"概率采样(5%),高风险标签,用户负反馈,规则兜底采样,人工指定"')
dv_sample.errorTitle = "抽样方式"
dv_sample.error = "请从下拉列表选择"
ws1.add_data_validation(dv_sample)
dv_sample.add(f"C2:C500")

dv_safety = DataValidation(type="list", formula1='"通过,1项遗漏,多项遗漏,严重风险"')
ws1.add_data_validation(dv_safety)
dv_safety.add(f"K2:K500")

dv_diverse = DataValidation(type="list", formula1='"优秀(≥85%),良好(70-84%),偏低(50-69%),差(<50%)"')
ws1.add_data_validation(dv_diverse)
dv_diverse.add(f"L2:L500")

dv_diff = DataValidation(type="list", formula1='"匹配,偏高,偏低,严重不匹配"')
ws1.add_data_validation(dv_diff)
dv_diff.add(f"M2:M500")

dv_explain = DataValidation(type="list", formula1='"清晰,部分缺失,不可解释"')
ws1.add_data_validation(dv_explain)
dv_explain.add(f"N2:N500")

dv_result = DataValidation(type="list", formula1='"推荐准确,轻微偏差,Bad Case"')
ws1.add_data_validation(dv_result)
dv_result.add(f"O2:O500")

dv_badtype = DataValidation(type="list", formula1='"安全禁忌遗漏,内容过于重复,难度不匹配,时长超限,规则逻辑矛盾,兜底内容不足,其他"')
ws1.add_data_validation(dv_badtype)
dv_badtype.add(f"P2:P500")

# 冻结首行
ws1.freeze_panes = "A2"


# ════════════════════════════════════════════════════════════════
# Sheet 2 ─ 评分标准与流程说明
# ════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("2-评分标准与流程")
ws2.sheet_properties.tabColor = "70AD47"

guide = [
    ["一、抽查评测标准流程", "", "", ""],
    ["步骤", "操作内容", "负责角色", "产出物"],
    ["1. 采样", "系统每日按 5% 概率 + 高风险标签 + 用户负反馈 + 兜底触发 四种方式自动采样", "系统自动", "抽查样本池（B31 列表页）"],
    ["2. 领取样本", "评测人在 B31 抽查池中领取\u201c待抽查\u201d样本（建议每人每日 \u226420 条）", "课程运营 / 健康专家", "样本状态变为\u201c评测中\u201d"],
    ["3. 查看问卷画像", "核对用户的问卷答案、健康标签、训练目标、器械条件等原始输入", "评测人", "理解用户需求"],
    ["4. 逐日审查课表", "逐天查看 AI 推荐的课程视频，检查安全性、难度、时长、多样性", "评测人", "逐日审查记录"],
    ["5. 查看推荐理由", "确认每条推荐附带的规则命中链与 AI 选课理由是否合理可解释", "评测人", "可解释性评估"],
    ["6. 四维打分", "从安全禁忌、内容多样性、难度匹配度、推荐可解释性四个维度独立打分", "评测人", "四维分数（填入Sheet1）"],
    ["7. 综合评定", "根据四维分数综合判定：推荐准确 / 轻微偏差 / Bad Case", "评测人", "综合评定结论"],
    ["8. Bad Case 归因", "如为 Bad Case，填写问题类型、具体描述、建议优化的 B12 规则 ID", "评测人", "Bad Case 归因记录"],
    ["9. 提交规则优化", "点击【一键跳转 B12】修改排课规则，或提交优化工单给算法团队", "课程运营 → 规则管理员", "B12 规则新版本"],
    ["10. 复验", "规则更新后，系统自动对同类标签用户重新抽样，验证修复效果", "系统 + 评测人", "修复确认 / 闭环"],
    ["", "", "", ""],
    ["二、四维评分标准", "", "", ""],
    ["评分维度", "评定标准", "推荐准确", "Bad Case 判定阈值"],
    ["① 安全禁忌校验", "检查 AI 是否正确排除了用户健康标签对应的禁忌动作（产后/腰突/经期等）", "通过 = 0 项遗漏", "≥1 项安全禁忌遗漏即为 Bad Case（一票否决）"],
    ["② 内容多样性", "7 天内同部位训练不连续超过 2 天；同一视频 7 天不重复", "≥70%", "<50% 判定为 Bad Case"],
    ["③ 难度匹配度", "AI 推荐的课程强度与用户问卷填写的运动基础、目标是否一致", "匹配", "严重不匹配（新手推荐高强度）为 Bad Case"],
    ["④ 推荐可解释性", "每节课附带的推荐理由（命中规则+选课原因）是否完整且合理", "清晰", "完全不可解释需标记并反馈算法团队"],
    ["", "", "", ""],
    ["三、综合评定规则", "", "", ""],
    ["综合评定", "判定条件", "", ""],
    ["推荐准确", "四项均达标（安全通过、多样性≥70%、难度匹配、理由清晰）", "", ""],
    ["轻微偏差", "安全通过但多样性/难度/解释性任一项略低于标准，不影响用户安全与体验底线", "", ""],
    ["Bad Case", "安全禁忌遗漏（一票否决）OR 多样性<50% OR 难度严重不匹配", "", ""],
    ["", "", "", ""],
    ["四、抽样方式说明", "", "", ""],
    ["抽样方式", "触发条件", "采样比例", "优先级"],
    ["概率采样(5%)", "每日全部实时生成课表中随机抽取", "5%", "常规"],
    ["高风险标签", "用户问卷包含产后/腹直肌分离/腰椎间盘/经期+大基数等敏感标签", "100% 自动入池", "高（安全优先）"],
    ["用户负反馈", "用户触发\u22653次\u201c太累降级\u201d或手动替换\u22652次课程", "100% 自动入池", "高（体验优先）"],
    ["规则兜底采样", "排课引擎未命中任何优先规则，退化为兜底池的课表", "100% 自动入池", "中"],
    ["人工指定", "运营人员根据业务需要手动指定特定用户进行深度评测", "按需", "按需"],
]

for r, row_data in enumerate(guide, 1):
    for c, val in enumerate(row_data, 1):
        cell = ws2.cell(row=r, column=c, value=val)
        cell.font = BODY_FONT
        cell.alignment = LEFT
        cell.border = BOX

section_rows = [1, 14, 21, 26]
for sr in section_rows:
    for c in range(1, 5):
        cell = ws2.cell(row=sr, column=c)
        cell.font = Font(name="微软雅黑", bold=True, size=12, color="FFFFFF")
        cell.fill = BLUE
        cell.alignment = CENTER

sub_header_rows = [2, 15, 22, 27]
for sr in sub_header_rows:
    for c in range(1, 5):
        cell = ws2.cell(row=sr, column=c)
        cell.font = SUB_FONT
        cell.fill = LTBLUE

ws2.column_dimensions["A"].width = 22
ws2.column_dimensions["B"].width = 60
ws2.column_dimensions["C"].width = 28
ws2.column_dimensions["D"].width = 36
ws2.freeze_panes = "A2"


# ════════════════════════════════════════════════════════════════
# Sheet 3 ─ Bad Case 归因分类参考
# ════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("3-Bad Case归因分类")
ws3.sheet_properties.tabColor = "FF0000"

bad_cases = [
    ["Bad Case 类型", "典型场景描述", "影响等级", "对应 B12 规则层", "建议优化方向", "示例"],
    ["安全禁忌遗漏", "用户标签含「产后」但被推荐了卷腹/仰卧起坐等腹直肌分离禁用动作", "P0 严重", "安全禁忌层", "在 B12 硬性排除中补全「产后→排除卷腹类」标签映射", "U-6102 产后+卷腹"],
    ["安全禁忌遗漏", "用户标签含「腰椎间盘突出」但被推荐深蹲/硬拉等高脊柱压力动作", "P0 严重", "安全禁忌层", "补全腰椎相关硬性排除标签", "U-3301 腰突+硬拉"],
    ["安全禁忌遗漏", "经期用户被推荐倒立/高冲击跳跃类动作", "P0 严重", "周期安全层", "在经期规则中增加「倒立/跳跃」排除", "U-4421 经期+跳跃"],
    ["内容过于重复", "连续 3 天推荐同系列/同部位训练视频", "P1 高", "内容多样性层", "提高多样性规则权重或优先级，使其不被难度/目标规则覆盖", "U-4320 连续3天哑铃"],
    ["内容过于重复", "7 天周期内同一视频出现 3 次以上", "P1 高", "内容多样性层", "加强「7日内同视频不重复」硬性约束", "U-5501 重复视频"],
    ["难度不匹配", "新手用户（运动基础=0）被推荐中高强度训练", "P1 高", "难度/时长层", "强化新手保护规则：运动基础=0时强度上限≤2", "U-7780 新手+高强度"],
    ["难度不匹配", "进阶用户长期被推荐入门级内容，无难度递进", "P2 中", "难度/时长层", "检查难度递进规则是否生效，增加周期内难度梯度约束", "U-2210 进阶+入门"],
    ["时长超限", "单日推荐课程总时长超过用户设定的目标时长上限", "P2 中", "难度/时长层", "在课表限制中严格执行时长上限约束", "U-3321 目标20min实际35min"],
    ["规则逻辑矛盾", "两条规则的硬性必须与硬性排除存在交集，导致候选集为空后退化兜底", "P1 高", "多层交叉", "在 B12 保存时增加跨规则矛盾检测", "R-04与R-07冲突"],
    ["兜底内容不足", "兜底池视频数量过少（<5条），用户课表出现大量重复兜底内容", "P2 中", "兜底层", "扩充兜底池内容；设置兜底池最小容量阈值告警", "兜底池仅3条视频"],
    ["其他", "未归入以上分类的推荐异常", "按实际判定", "按实际判定", "描述具体问题并提交算法团队分析", ""],
]

for r, row_data in enumerate(bad_cases, 1):
    for c, val in enumerate(row_data, 1):
        cell = ws3.cell(row=r, column=c, value=val)
        cell.font = BODY_FONT
        cell.alignment = LEFT
        cell.border = BOX
style_header(ws3, 1, 6)
style_rows(ws3, 2, len(bad_cases), 6)

# 按影响等级高亮
for r in range(2, len(bad_cases) + 1):
    level = ws3.cell(row=r, column=3).value
    if level and "P0" in str(level):
        for c in range(1, 7):
            ws3.cell(row=r, column=c).fill = RED
    elif level and "P1" in str(level):
        for c in range(1, 7):
            ws3.cell(row=r, column=c).fill = YELLOW

ws3.column_dimensions["A"].width = 20
ws3.column_dimensions["B"].width = 50
ws3.column_dimensions["C"].width = 14
ws3.column_dimensions["D"].width = 20
ws3.column_dimensions["E"].width = 48
ws3.column_dimensions["F"].width = 26
ws3.freeze_panes = "A2"


# ════════════════════════════════════════════════════════════════
# Sheet 4 ─ 周报汇总模板
# ════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("4-周报汇总模板")
ws4.sheet_properties.tabColor = "FFC000"

report = [
    ["AI 排课推荐准确度 · 周报", "", "", "", "", "", ""],
    ["统计周期", "2026-08-12 ~ 2026-08-18", "", "", "", "", ""],
    ["", "", "", "", "", "", ""],
    ["一、大盘数据", "", "", "", "", "", ""],
    ["指标", "本周", "上周", "环比", "目标值", "达标", "备注"],
    ["实时生成课表总数", "8,960", "8,120", "+10.3%", "-", "-", ""],
    ["抽查样本总数", "448", "406", "+10.3%", "≥5%采样率", "达标", ""],
    ["推荐准确率", "96.4%", "95.1%", "+1.3pp", "≥95%", "达标", ""],
    ["Bad Case 数量", "16", "20", "-20%", "≤25", "达标", ""],
    ["Bad Case 率", "3.6%", "4.9%", "-1.3pp", "≤5%", "达标", ""],
    ["安全禁忌 Bad Case", "2", "4", "-50%", "0 (目标)", "未达标", "已全部修复规则"],
    ["已联动优化规则数", "5", "3", "+67%", "-", "-", ""],
    ["规则修复后复验通过", "4/5", "2/3", "", "100%", "80%", "剩余1条下周复验"],
    ["", "", "", "", "", "", ""],
    ["二、Bad Case 分布", "", "", "", "", "", ""],
    ["Bad Case 类型", "本周数量", "占比", "本周新增规则优化", "闭环状态", "", ""],
    ["安全禁忌遗漏", "2", "12.5%", "R-07 产后安全禁忌 v2", "已修复", "", ""],
    ["内容过于重复", "6", "37.5%", "R-11 多样性权重提升", "已修复", "", ""],
    ["难度不匹配", "4", "25.0%", "R-05 新手保护强度上限", "评测中", "", ""],
    ["时长超限", "2", "12.5%", "R-06 时长硬约束", "已修复", "", ""],
    ["规则逻辑矛盾", "1", "6.25%", "B12 跨规则冲突检测", "开发中", "", ""],
    ["兜底内容不足", "1", "6.25%", "扩充兜底池+阈值告警", "计划中", "", ""],
    ["", "", "", "", "", "", ""],
    ["三、本周重点优化动作", "", "", "", "", "", ""],
    ["优化序号", "问题来源样本", "问题摘要", "优化动作", "涉及规则", "状态", "负责人"],
    ["OPT-01", "U-6102", "产后用户被推荐卷腹", "B12 新增产后硬性排除标签映射", "R-07 v2", "已上线", "李四"],
    ["OPT-02", "U-4320", "连续3天同系列哑铃视频", "提升 R-11 多样性权重至第3优先级", "R-11 v3", "已上线", "张三"],
    ["OPT-03", "U-7780", "新手推荐高强度", "新手保护层增加强度上限≤2", "R-05 v6", "测试中", "王五"],
    ["", "", "", "", "", "", ""],
    ["填表人：", "", "审核人：", "", "日期：", "", ""],
]

for r, row_data in enumerate(report, 1):
    for c, val in enumerate(row_data, 1):
        cell = ws4.cell(row=r, column=c, value=val)
        cell.font = BODY_FONT
        cell.alignment = LEFT
        cell.border = BOX

# 标题行样式
title_rows = [1, 4, 15, 24]
for tr in title_rows:
    for c in range(1, 8):
        cell = ws4.cell(row=tr, column=c)
        cell.font = Font(name="微软雅黑", bold=True, size=12 if tr == 1 else 11, color="FFFFFF")
        cell.fill = BLUE
        cell.alignment = CENTER

sub_rows = [5, 16, 25]
for sr in sub_rows:
    for c in range(1, 8):
        cell = ws4.cell(row=sr, column=c)
        cell.font = SUB_FONT
        cell.fill = LTBLUE

ws4.column_dimensions["A"].width = 20
ws4.column_dimensions["B"].width = 18
ws4.column_dimensions["C"].width = 18
ws4.column_dimensions["D"].width = 36
ws4.column_dimensions["E"].width = 24
ws4.column_dimensions["F"].width = 14
ws4.column_dimensions["G"].width = 14
ws4.freeze_panes = "A2"

# 合并标题
ws4.merge_cells("A1:G1")
ws4.merge_cells("A4:G4")
ws4.merge_cells("A15:G15")
ws4.merge_cells("A24:G24")


# ════════════════════════════════════════════════════════════════
# Sheet 5 ─ 逐日课表审查明细（单个用户的深度评测）
# ════════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("5-逐日课表审查明细")
ws5.sheet_properties.tabColor = "7030A0"

ws5.cell(row=1, column=1, value="用户样本 ID")
ws5.cell(row=1, column=2, value="U-8921")
ws5.cell(row=2, column=1, value="问卷核心标签")
ws5.cell(row=2, column=2, value="黄体期 / 大基数 / 目标舒缓 / 无器械")
ws5.cell(row=3, column=1, value="AI 推荐方案")
ws5.cell(row=3, column=2, value="舒缓减压·7天")
ws5.cell(row=4, column=1, value="命中规则链")
ws5.cell(row=4, column=2, value="R-04 周期安全层 → R-09 低冲击偏好 → R-11 多样性")
ws5.cell(row=5, column=1, value="评测人")
ws5.cell(row=5, column=2, value="（填写）")
ws5.cell(row=6, column=1, value="评测日期")
ws5.cell(row=6, column=2, value="（填写）")

for r in range(1, 7):
    for c in range(1, 3):
        cell = ws5.cell(row=r, column=c)
        cell.font = BODY_FONT if c == 2 else SUB_FONT
        cell.border = BOX
        cell.alignment = LEFT
        if c == 1:
            cell.fill = LTBLUE
ws5.column_dimensions["A"].width = 18
ws5.column_dimensions["B"].width = 50

cols5 = [
    ("天", 6), ("推荐视频名称", 28), ("时长(min)", 12), ("强度等级", 12),
    ("命中规则", 22), ("AI推荐理由", 36), ("安全校验", 12), ("多样性", 12),
    ("难度合理", 12), ("评测结论", 14), ("问题备注", 30),
]
for c, (name, w) in enumerate(cols5, 1):
    ws5.cell(row=8, column=c, value=name)
    ws5.column_dimensions[get_column_letter(c)].width = w
style_header(ws5, 8, len(cols5), fill=PatternFill("solid", fgColor="7030A0"))

day_examples = [
    ["D1", "呼吸与全身放松", 15, "低", "R-04", "黄体期首日轻度放松优先", "通过", "—", "匹配", "准确", ""],
    ["D2", "温和全身拉伸", 18, "低", "R-04+R-09", "延续低冲击基调", "通过", "良好", "匹配", "准确", ""],
    ["D3", "骨盆稳定修护", 20, "低-中", "R-09", "避免膝关节冲击", "通过", "良好", "匹配", "准确", ""],
    ["D4", "核心温和激活", 22, "低-中", "R-09", "温和核心训练", "通过", "与D3部位近", "匹配", "轻微偏差", "动作与D3部分重叠"],
    ["D5", "休息日", 0, "—", "—", "每周1个休息日规则", "—", "—", "—", "准确", ""],
    ["D6", "下肢舒缓放松", 18, "低", "R-09", "切换部位增加多样性", "通过", "良好", "匹配", "准确", ""],
    ["D7", "深度全身拉伸", 18, "低", "R-12", "7天小周期收尾", "通过", "良好", "匹配", "准确", ""],
]
for r, row_data in enumerate(day_examples, 9):
    for c, val in enumerate(row_data, 1):
        ws5.cell(row=r, column=c, value=val)
style_rows(ws5, 9, 15, len(cols5))

# 下拉验证
dv5_safe = DataValidation(type="list", formula1='"通过,遗漏,—"')
ws5.add_data_validation(dv5_safe)
dv5_safe.add("G9:G200")

dv5_result = DataValidation(type="list", formula1='"准确,轻微偏差,Bad Case,—"')
ws5.add_data_validation(dv5_result)
dv5_result.add("J9:J200")

ws5.freeze_panes = "A9"


# ════════════════════════════════════════════════════════════════
# 保存
# ════════════════════════════════════════════════════════════════
out_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), "docs")
os.makedirs(out_dir, exist_ok=True)
out_path = os.path.join(out_dir, "AI排课推荐抽查评测模板.xlsx")
wb.save(out_path)
print(f"✅ 模板已生成：{out_path}")
